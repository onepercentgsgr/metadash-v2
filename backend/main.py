from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Header, Request
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.responses import Response as StarletteResponse
from sqlalchemy.orm import Session
from datetime import datetime, timedelta
from typing import Optional, List, Any, Dict
from pydantic import BaseModel
import jwt
from passlib.context import CryptContext
import os
from dotenv import load_dotenv
import json
import openpyxl
import io
import logging
from slowapi import Limiter
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from fastapi.responses import JSONResponse

from database import engine, get_db, Base
from models import User, TenantConfig, Subscription, AgentLog, FinancialRecord, ShopifyOrder, AutonomousActionLog
from agents.shared_memory import AgentMemory  # registers table with Base.metadata
from security_middleware import SecurityHeadersMiddleware
import config

logger = logging.getLogger(__name__)

load_dotenv()

# Import payment routes
try:
    from payment_routes import router as payment_router
    PAYMENTS_AVAILABLE = True
except ImportError:
    PAYMENTS_AVAILABLE = False

# Import v3.5 routes
try:
    from v35_routes import router as v35_router
    V35_AVAILABLE = True
except ImportError:
    V35_AVAILABLE = False

# Initialize database
Base.metadata.create_all(bind=engine)

# Create FastAPI app
app = FastAPI(title="MetaDash API", version="3.5.0")

# Setup rate limiting
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter

@app.exception_handler(RateLimitExceeded)
async def rate_limit_exceeded_handler(request: Request, exc: RateLimitExceeded):
    return JSONResponse(
        status_code=429,
        content={"detail": "Rate limit exceeded. Please try again later."}
    )

# ── CORS Configuration ──
# Secure allowlist of origins - reject everything else
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "http://localhost:3000").split(",")
ALLOWED_ORIGINS = [origin.strip() for origin in ALLOWED_ORIGINS]

class SecureCORSMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        origin = request.headers.get("origin", "")

        # Only respond to whitelisted origins
        allowed_origin = origin if origin in ALLOWED_ORIGINS else None

        # Preflight OPTIONS → responder inmediatamente con CORS headers
        if request.method == "OPTIONS":
            response = StarletteResponse(status_code=200)
            if allowed_origin:
                response.headers["Access-Control-Allow-Origin"] = allowed_origin
                response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS, PATCH"
                response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization, X-Requested-With"
                response.headers["Access-Control-Allow-Credentials"] = "true"
                response.headers["Access-Control-Max-Age"] = "3600"
            return response

        # Request normal → ejecutar y agregar CORS headers a la respuesta
        response = await call_next(request)

        if allowed_origin:
            response.headers["Access-Control-Allow-Origin"] = allowed_origin
            response.headers["Access-Control-Allow-Credentials"] = "true"
            response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS, PATCH"
            response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization, X-Requested-With"

        return response


app.add_middleware(SecureCORSMiddleware)
app.add_middleware(SecurityHeadersMiddleware)

# Include payment routes
if PAYMENTS_AVAILABLE:
    app.include_router(payment_router)

# Include v3.5 routes
if V35_AVAILABLE:
    app.include_router(v35_router)

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Request/Response Models
class UserRegister(BaseModel):
    email: str
    password: str
    name: str


class UserLogin(BaseModel):
    email: str
    password: str


class TokenResponse(BaseModel):
    access_token: str
    token_type: str
    user_id: int
    email: str
    name: str
    role: str
    subscription_plan: str
    subscription_status: str
    trial_end: Optional[datetime] = None


class UserResponse(BaseModel):
    id: int
    email: str
    name: str
    role: str
    is_active: bool
    created_at: datetime
    subscription_plan: Optional[str] = None
    subscription_status: Optional[str] = None
    trial_end: Optional[datetime] = None

    class Config:
        from_attributes = True


class TenantConfigUpdate(BaseModel):
    meta_access_token: Optional[str] = None
    meta_ad_account_id: Optional[str] = None
    meta_app_id: Optional[str] = None
    meta_app_secret: Optional[str] = None
    anthropic_api_key: Optional[str] = None
    hf_api_key: Optional[str] = None
    negocio_info: Optional[str] = None
    landing_page_url: Optional[str] = None
    shopify_store_url: Optional[str] = None
    shopify_webhook_secret: Optional[str] = None
    mercadopago_access_token: Optional[str] = None


class TenantConfigResponse(BaseModel):
    id: int
    user_id: int
    meta_access_token: Optional[str] = None
    meta_ad_account_id: Optional[str] = None
    meta_app_id: Optional[str] = None
    meta_app_secret: Optional[str] = None
    anthropic_api_key: Optional[str] = None
    hf_api_key: Optional[str] = None
    negocio_info: Optional[str] = None
    landing_page_url: Optional[str] = None
    shopify_store_url: Optional[str] = None
    shopify_webhook_secret: Optional[str] = None
    mercadopago_access_token: Optional[str] = None

    class Config:
        from_attributes = True


class AgentRequest(BaseModel):
    prompt: str = ""
    context: Optional[dict] = None


class AgentResponse(BaseModel):
    result: Any
    agent: str


class AdminUserResponse(BaseModel):
    id: int
    email: str
    name: str
    is_active: bool
    created_at: datetime
    subscription_plan: str
    subscription_status: str
    trial_end: Optional[datetime] = None

    class Config:
        from_attributes = True


class AdminStats(BaseModel):
    total_users: int
    active_trials: int
    paid_users: int
    revenue_ars: float
    revenue_usd: float
    trial_expiring_soon: int
    expired_trials: int
    mrr: float
    avg_arpu: float
    videos_generated: int
    agents_runs: int
    churn_rate: float


class FinancialRecordCreate(BaseModel):
    periodo: str
    ingresos: Optional[int] = None
    costos: Optional[int] = None
    ad_spend: Optional[int] = None
    devoluciones: Optional[int] = None
    ordenes: Optional[int] = None


class FinancialRecordResponse(BaseModel):
    id: int
    user_id: int
    periodo: str
    ingresos: Optional[int]
    costos: Optional[int]
    ad_spend: Optional[int]
    devoluciones: Optional[int]
    ordenes: Optional[int]
    created_at: datetime

    class Config:
        from_attributes = True


class AdminUserUpdate(BaseModel):
    role: Optional[str] = None
    is_active: Optional[bool] = None


class AdminSetPlanRequest(BaseModel):
    plan: str


class SubscriptionResponse(BaseModel):
    id: int
    user_id: int
    plan: str
    status: str
    trial_start: Optional[datetime] = None
    trial_end: Optional[datetime] = None
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class SubscriptionUpdate(BaseModel):
    plan: Optional[str] = None
    status: Optional[str] = None
    trial_start: Optional[datetime] = None
    trial_end: Optional[datetime] = None


class AgentRunRequest(BaseModel):
    agent_type: str
    input: str
    context: Optional[dict] = None


class ShopifyOrderResponse(BaseModel):
    id: int
    user_id: int
    order_id: str
    order_number: str
    email: str
    total_price: str
    subtotal_price: str
    total_tax: str
    currency: str
    financial_status: str
    fulfillment_status: str
    customer_first_name: Optional[str]
    customer_last_name: Optional[str]
    created_at: datetime

    class Config:
        from_attributes = True


# ── Playbook Nivel Dios Models ──

class PlaybookCopywriterRequest(BaseModel):
    nicho: str
    audience: str
    pain_point: str
    mechanism_name: str
    mechanism_description: str
    price: float = 17
    bonos: List[str] = []
    tone: str = "cercano, humilde y motivador"


class PlaybookDesignRequest(BaseModel):
    product_type: str  # "PDF", "VIDEO", "TEMPLATE", "CURSO"
    product_name: str
    nicho: str
    price: float = 17
    headline: str = ""
    pain_point: str = ""
    emotion_target: str = "trust"


class PlaybookSocialMediaRequest(BaseModel):
    nicho: str
    pain_point: str
    mechanism_name: str
    audience: str
    hook: str = ""
    angle_type: str = "pain"
    format_type: str = "9x16"
    winning_creative_stats: Optional[Dict[str, Any]] = None
    current_budget: float = 10


# Utility Functions
def hash_password(password: str) -> str:
    # bcrypt has a 72-byte limit, truncate to prevent ValueError
    return pwd_context.hash(password[:72])


def verify_password(plain_password: str, hashed_password: str) -> bool:
    # bcrypt has a 72-byte limit, truncate to match hash_password
    return pwd_context.verify(plain_password[:72], hashed_password)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None) -> str:
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(hours=24)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, config.SECRET_KEY, algorithm=config.ALGORITHM)
    return encoded_jwt


def verify_token(token: str) -> dict:
    try:
        logger.debug(f"verify_token: Decoding token, length={len(token)}, starts_with={token[:20] if len(token) > 20 else token}")
        payload = jwt.decode(token, config.SECRET_KEY, algorithms=[config.ALGORITHM])
        sub = payload.get("sub")
        if sub is None:
            logger.error("verify_token: No 'sub' field in JWT payload")
            raise HTTPException(status_code=401, detail="Invalid token")
        # Convert sub to int (it's stored as string in JWT)
        try:
            user_id: int = int(sub)
        except (ValueError, TypeError):
            logger.error(f"verify_token: Invalid user_id format: {sub}")
            raise HTTPException(status_code=401, detail="Invalid token")
        logger.debug(f"verify_token: Token verified for user_id={user_id}")
        return {"user_id": user_id}
    except jwt.ExpiredSignatureError:
        logger.warning("verify_token: Token expired")
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError as e:
        logger.error(f"verify_token: Invalid token - {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=401, detail="Invalid token")


def get_current_user(
    token: Optional[str] = None,
    db: Session = Depends(get_db),
) -> User:
    if not token:
        raise HTTPException(status_code=401, detail="No token provided")
    
    payload = verify_token(token)
    user_id = payload.get("user_id")
    
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    return user


def get_current_user_from_header(
    authorization: Optional[str] = None,
    db: Session = Depends(get_db),
) -> User:
    if not authorization:
        logger.error("get_current_user_from_header: No authorization header received")
        raise HTTPException(status_code=401, detail="No authorization header")

    try:
        logger.debug(f"get_current_user_from_header: Authorization header present, length={len(authorization)}")
        token = authorization.replace("Bearer ", "")
        if not token or len(token) < 10:
            logger.error(f"get_current_user_from_header: Token extraction failed or too short: {len(token) if token else 0}")
            raise HTTPException(status_code=401, detail="Invalid token format")

        payload = verify_token(token)
        user_id = payload.get("user_id")

        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            logger.error(f"get_current_user_from_header: User not found for id={user_id}")
            raise HTTPException(status_code=404, detail="User not found")

        logger.debug(f"get_current_user_from_header: Successfully authenticated user={user.email}")
        return user
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"get_current_user_from_header: Unexpected error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=401, detail="Authentication failed")


def get_tenant_config(user_id: int, db: Session) -> TenantConfig:
    config_obj = db.query(TenantConfig).filter(TenantConfig.user_id == user_id).first()
    if not config_obj:
        raise HTTPException(status_code=404, detail="Tenant config not found")
    return config_obj


def get_active_subscription(user_id: int, db: Session) -> Subscription:
    subscription = db.query(Subscription).filter(
        Subscription.user_id == user_id,
        Subscription.status == "active"
    ).order_by(Subscription.created_at.desc()).first()
    
    if not subscription:
        raise HTTPException(status_code=403, detail="No active subscription")
    
    # Check if trial has expired
    if subscription.plan == "trial" and subscription.trial_end:
        if datetime.utcnow() > subscription.trial_end:
            subscription.status = "expired"
            db.commit()
            raise HTTPException(status_code=403, detail="Trial expired")
    
    return subscription


async def check_subscription(
    authorization: Optional[str] = None,
    db: Session = Depends(get_db),
) -> User:
    user = get_current_user_from_header(authorization, db)
    
    # Admin users bypass subscription check
    if user.role == "admin":
        return user
    
    # Check active subscription
    subscription = get_active_subscription(user.id, db)
    
    return user


# Startup Event
@app.on_event("startup")
def startup_event():
    db = next(get_db())
    try:
        # Create admin user if not exists
        admin_email = config.ADMIN_EMAIL
        admin_password = config.ADMIN_PASSWORD
        
        admin = db.query(User).filter(User.email == admin_email).first()
        if not admin:
            admin = User(
                email=admin_email,
                hashed_password=hash_password(admin_password),
                name="Admin",
                role="admin",
                is_active=True,
            )
            db.add(admin)
            db.commit()
            db.refresh(admin)

            # Create admin tenant config
            tenant_config = TenantConfig(user_id=admin.id)
            db.add(tenant_config)

            # Create admin trial subscription
            trial_end = datetime.utcnow() + timedelta(days=365)
            subscription = Subscription(
                user_id=admin.id,
                plan="enterprise",
                status="active",
                trial_start=datetime.utcnow(),
                trial_end=trial_end,
            )
            db.add(subscription)
            db.commit()
        elif admin.role != "admin":
            # If admin email user exists but isn't admin yet, upgrade them
            admin.role = "admin"
            admin.is_active = True
            db.commit()
            logger.info(f"Upgraded {admin_email} to admin role")

        # Ensure admin has enterprise plan (not trial)
        admin_sub = db.query(Subscription).filter(
            Subscription.user_id == admin.id
        ).order_by(Subscription.created_at.desc()).first()
        if admin_sub and admin_sub.plan == "trial":
            admin_sub.plan = "enterprise"
            admin_sub.status = "active"
            admin_sub.trial_end = datetime.utcnow() + timedelta(days=365)
            db.commit()
    finally:
        db.close()

    # Start autonomous scheduler
    try:
        from jobs.scheduler import schedule_autonomous_audits
        schedule_autonomous_audits()
    except ImportError:
        logger.warning("Jobs module not available, autonomous audits disabled")
    except Exception as e:
        logger.error(f"Error starting scheduler: {str(e)}")


# Shutdown Event
@app.on_event("shutdown")
def shutdown_event():
    try:
        from jobs.scheduler import stop_scheduler
        stop_scheduler()
    except:
        pass

# Health Endpoints
@app.get("/")
@app.get("/health")
async def health_check():
    return {
        "status": "ok",
        "service": "MetaDash API",
        "version": "3.5.0",
    }


# Auth Endpoints
@app.post("/auth/register", response_model=TokenResponse)
async def register(
    request: UserRegister,
    db: Session = Depends(get_db),
):
    # Check if user exists
    existing_user = db.query(User).filter(User.email == request.email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create user
    user = User(
        email=request.email,
        hashed_password=hash_password(request.password),
        name=request.name,
        role="client",
        is_active=True,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    
    # Create tenant config
    tenant_config = TenantConfig(user_id=user.id)
    db.add(tenant_config)
    
    # Create trial subscription (7 days)
    trial_end = datetime.utcnow() + timedelta(days=7)
    subscription = Subscription(
        user_id=user.id,
        plan="trial",
        status="active",
        trial_start=datetime.utcnow(),
        trial_end=trial_end,
    )
    db.add(subscription)
    db.commit()
    
    # Generate token
    access_token = create_access_token(
        data={"sub": str(user.id)},
        expires_delta=timedelta(hours=24),
    )
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user_id": user.id,
        "email": user.email,
        "name": user.name,
        "role": user.role,
        "subscription_plan": subscription.plan,
        "subscription_status": subscription.status,
        "trial_end": subscription.trial_end,
    }


@app.post("/auth/login", response_model=TokenResponse)
async def login(
    request: UserLogin,
    db: Session = Depends(get_db),
):
    # Find user
    user = db.query(User).filter(User.email == request.email).first()
    if not user or not verify_password(request.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Get latest subscription
    subscription = db.query(Subscription).filter(
        Subscription.user_id == user.id,
    ).order_by(Subscription.created_at.desc()).first()
    
    if not subscription:
        raise HTTPException(status_code=500, detail="No subscription found")
    
    # Generate token
    access_token = create_access_token(
        data={"sub": str(user.id)},
        expires_delta=timedelta(hours=24),
    )
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user_id": user.id,
        "email": user.email,
        "name": user.name,
        "role": user.role,
        "subscription_plan": subscription.plan,
        "subscription_status": subscription.status,
        "trial_end": subscription.trial_end,
    }


@app.get("/auth/me", response_model=UserResponse)
async def get_me(
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = get_current_user_from_header(authorization, db)
    
    subscription = db.query(Subscription).filter(
        Subscription.user_id == user.id,
    ).order_by(Subscription.created_at.desc()).first()
    
    return UserResponse(
        id=user.id,
        email=user.email,
        name=user.name,
        role=user.role,
        is_active=user.is_active,
        created_at=user.created_at,
        subscription_plan=subscription.plan if subscription else None,
        subscription_status=subscription.status if subscription else None,
        trial_end=subscription.trial_end if subscription else None,
    )


# Config Endpoints
@app.get("/config", response_model=TenantConfigResponse)
async def get_config(
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = get_current_user_from_header(authorization, db)
    config_obj = get_tenant_config(user.id, db)
    
    # Mask sensitive fields
    response = TenantConfigResponse.model_validate(config_obj)
    if response.meta_access_token:
        response.meta_access_token = response.meta_access_token[-10:] if len(response.meta_access_token) > 10 else "***"
    if response.anthropic_api_key:
        response.anthropic_api_key = response.anthropic_api_key[-10:] if len(response.anthropic_api_key) > 10 else "***"
    if response.hf_api_key:
        response.hf_api_key = response.hf_api_key[-10:] if len(response.hf_api_key) > 10 else "***"
    
    return response


@app.post("/config", response_model=TenantConfigResponse)
async def update_config(
    request: TenantConfigUpdate,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = get_current_user_from_header(authorization, db)
    config_obj = get_tenant_config(user.id, db)

    update_data = request.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(config_obj, field, value)

    db.commit()
    db.refresh(config_obj)

    return TenantConfigResponse.model_validate(config_obj)


@app.get("/tenant-config", response_model=TenantConfigResponse)
async def get_tenant_config_endpoint(
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = get_current_user_from_header(authorization, db)
    config_obj = get_tenant_config(user.id, db)

    # Mask sensitive fields
    response = TenantConfigResponse.model_validate(config_obj)
    if response.meta_access_token:
        response.meta_access_token = response.meta_access_token[-10:] if len(response.meta_access_token) > 10 else "***"
    if response.anthropic_api_key:
        response.anthropic_api_key = response.anthropic_api_key[-10:] if len(response.anthropic_api_key) > 10 else "***"
    if response.hf_api_key:
        response.hf_api_key = response.hf_api_key[-10:] if len(response.hf_api_key) > 10 else "***"

    return response


@app.put("/tenant-config", response_model=TenantConfigResponse)
async def update_tenant_config_endpoint(
    request: TenantConfigUpdate,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = get_current_user_from_header(authorization, db)
    config_obj = get_tenant_config(user.id, db)

    update_data = request.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(config_obj, field, value)

    db.commit()
    db.refresh(config_obj)

    return TenantConfigResponse.model_validate(config_obj)


# Subscription Endpoints
@app.get("/subscription")
async def get_subscription(
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = get_current_user_from_header(authorization, db)
    subscription = db.query(Subscription).filter(
        Subscription.user_id == user.id,
    ).order_by(Subscription.created_at.desc()).first()
    
    if not subscription:
        raise HTTPException(status_code=404, detail="No subscription found")
    
    return {
        "id": subscription.id,
        "plan": subscription.plan,
        "status": subscription.status,
        "trial_start": subscription.trial_start,
        "trial_end": subscription.trial_end,
        "created_at": subscription.created_at,
        "updated_at": subscription.updated_at,
    }


# Admin Endpoints
@app.get("/admin/users", response_model=List[AdminUserResponse])
async def list_users(
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = get_current_user_from_header(authorization, db)
    
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    users = db.query(User).all()
    result = []
    
    for u in users:
        subscription = db.query(Subscription).filter(
            Subscription.user_id == u.id,
        ).order_by(Subscription.created_at.desc()).first()
        
        result.append(AdminUserResponse(
            id=u.id,
            email=u.email,
            name=u.name,
            is_active=u.is_active,
            created_at=u.created_at,
            subscription_plan=subscription.plan if subscription else "none",
            subscription_status=subscription.status if subscription else "none",
            trial_end=subscription.trial_end if subscription else None,
        ))
    
    return result


@app.post("/admin/users/{user_id}/toggle")
async def toggle_user(
    user_id: int,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    admin = get_current_user_from_header(authorization, db)
    
    if admin.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    user.is_active = not user.is_active
    db.commit()
    
    return {"id": user.id, "is_active": user.is_active}


@app.post("/admin/users/{user_id}/extend-trial")
async def extend_trial(
    user_id: int,
    days: int = 7,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    admin = get_current_user_from_header(authorization, db)
    
    if admin.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    subscription = db.query(Subscription).filter(
        Subscription.user_id == user_id,
        Subscription.plan == "trial",
    ).order_by(Subscription.created_at.desc()).first()
    
    if not subscription:
        raise HTTPException(status_code=404, detail="Trial subscription not found")
    
    subscription.trial_end = subscription.trial_end + timedelta(days=days)
    db.commit()
    
    return {
        "user_id": user_id,
        "new_trial_end": subscription.trial_end,
    }


@app.post("/admin/users/{user_id}/set-plan")
async def set_plan(
    user_id: int,
    request: AdminSetPlanRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    admin = get_current_user_from_header(authorization, db)
    plan = request.plan

    if admin.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

    if plan not in ["trial", "starter", "pro", "enterprise"]:
        raise HTTPException(status_code=400, detail="Invalid plan")
    
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    subscription = db.query(Subscription).filter(
        Subscription.user_id == user_id,
    ).order_by(Subscription.created_at.desc()).first()
    
    if subscription:
        subscription.plan = plan
        subscription.status = "active"
        if plan == "trial":
            subscription.trial_start = datetime.utcnow()
            subscription.trial_end = datetime.utcnow() + timedelta(days=7)
        db.commit()
    else:
        new_subscription = Subscription(
            user_id=user_id,
            plan=plan,
            status="active",
            trial_start=datetime.utcnow() if plan == "trial" else None,
            trial_end=datetime.utcnow() + timedelta(days=7) if plan == "trial" else None,
        )
        db.add(new_subscription)
        db.commit()
    
    return {
        "user_id": user_id,
        "plan": plan,
        "status": "active",
    }


@app.get("/admin/stats", response_model=AdminStats)
async def get_admin_stats(
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    admin = get_current_user_from_header(authorization, db)
    if admin.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

    plan_prices_ars = {"starter": 19900, "pro": 29900, "enterprise": 79900}
    usd_rate = 1100

    total_users = db.query(User).count()
    active_trials = db.query(Subscription).filter(
        Subscription.plan == "trial",
        Subscription.status == "active",
    ).count()

    from datetime import timedelta
    soon_expire = db.query(Subscription).filter(
        Subscription.plan == "trial",
        Subscription.status == "active",
        Subscription.trial_end.between(datetime.utcnow(), datetime.utcnow() + timedelta(days=3))
    ).count()

    expired_trials = db.query(Subscription).filter(
        Subscription.status == "expired"
    ).count()

    paid_users = db.query(Subscription).filter(
        Subscription.plan.in_(["starter", "pro", "enterprise"]),
        Subscription.status == "active",
    ).count()

    paid_subscriptions = db.query(Subscription).filter(
        Subscription.plan.in_(["starter", "pro", "enterprise"]),
        Subscription.status == "active",
    ).all()

    revenue_ars = sum(plan_prices_ars.get(sub.plan, 0) for sub in paid_subscriptions)
    revenue_usd = revenue_ars / usd_rate
    mrr = revenue_ars
    arpu = (revenue_ars / paid_users) if paid_users > 0 else 0

    try:
        from models import GeneratedVideo
        videos_count = db.query(GeneratedVideo).count()
    except Exception:
        videos_count = 0
    try:
        agents_count = db.query(AgentLog).count()
    except Exception:
        agents_count = 0
    churn_rate = (expired_trials / (paid_users + expired_trials) * 100) if (paid_users + expired_trials) > 0 else 0

    return AdminStats(
        total_users=total_users,
        active_trials=active_trials,
        paid_users=paid_users,
        revenue_ars=float(revenue_ars),
        revenue_usd=float(revenue_usd),
        trial_expiring_soon=soon_expire,
        expired_trials=expired_trials,
        mrr=float(mrr),
        avg_arpu=float(arpu),
        videos_generated=videos_count,
        agents_runs=agents_count,
        churn_rate=float(churn_rate),
    )


@app.get("/admin/analytics/detailed")
async def get_detailed_analytics(
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    admin = get_current_user_from_header(authorization, db)
    if admin.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

    plan_prices_ars = {"starter": 19900, "pro": 29900, "enterprise": 79900}
    usd_rate = 1100

    plans_data = {}
    for plan in ["starter", "pro", "enterprise"]:
        subs = db.query(Subscription).filter(
            Subscription.plan == plan,
            Subscription.status == "active"
        ).all()
        revenue = sum(plan_prices_ars.get(plan, 0) for _ in subs)
        plans_data[plan] = {
            "count": len(subs),
            "revenue_ars": revenue,
            "revenue_usd": revenue / usd_rate,
        }

    from sqlalchemy import func
    try:
        from models import GeneratedVideo
        top_users = db.query(
            User.id, User.email, User.name,
            func.count(GeneratedVideo.id).label("videos_count")
        ).outerjoin(GeneratedVideo).filter(
            User.is_active == True
        ).group_by(User.id, User.email, User.name).order_by(
            func.count(GeneratedVideo.id).desc()
        ).limit(10).all()

        top_users_list = [
            {"user_id": u[0], "email": u[1], "name": u[2], "videos_generated": u[3] or 0}
            for u in top_users
        ]
        total_videos = db.query(GeneratedVideo).count()
    except Exception:
        top_users_list = []
        total_videos = 0

    try:
        total_agent_runs = db.query(AgentLog).count()
        total_agents = db.query(AgentLog).filter(AgentLog.status == "success").count()
    except Exception:
        total_agent_runs = 0
        total_agents = 0

    return {
        "plans_breakdown": plans_data,
        "top_users": top_users_list,
        "usage_metrics": {
            "total_videos": total_videos,
            "total_agent_runs": total_agent_runs,
            "successful_runs": total_agents,
        },
        "timestamp": datetime.utcnow().isoformat(),
    }


@app.put("/admin/users/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: int,
    request: AdminUserUpdate,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    admin = get_current_user_from_header(authorization, db)

    if admin.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    if request.role is not None:
        user.role = request.role
    if request.is_active is not None:
        user.is_active = request.is_active

    db.commit()
    db.refresh(user)

    subscription = db.query(Subscription).filter(
        Subscription.user_id == user.id,
    ).order_by(Subscription.created_at.desc()).first()

    return UserResponse(
        id=user.id,
        email=user.email,
        name=user.name,
        role=user.role,
        is_active=user.is_active,
        created_at=user.created_at,
        subscription_plan=subscription.plan if subscription else None,
        subscription_status=subscription.status if subscription else None,
        trial_end=subscription.trial_end if subscription else None,
    )


@app.delete("/admin/users/{user_id}")
async def delete_user(
    user_id: int,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    admin = get_current_user_from_header(authorization, db)

    if admin.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")


    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Soft delete - deactivate the user
    user.is_active = False
    db.commit()

    return {"id": user.id, "is_active": user.is_active, "message": "User deactivated"}


@app.get("/admin/subscriptions", response_model=List[SubscriptionResponse])
async def list_subscriptions(
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    admin = get_current_user_from_header(authorization, db)

    if admin.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

    subscriptions = db.query(Subscription).all()
    return [SubscriptionResponse.model_validate(sub) for sub in subscriptions]


@app.put("/admin/subscriptions/{sub_id}", response_model=SubscriptionResponse)
async def update_subscription(
    sub_id: int,
    request: SubscriptionUpdate,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    admin = get_current_user_from_header(authorization, db)

    if admin.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

    subscription = db.query(Subscription).filter(Subscription.id == sub_id).first()
    if not subscription:
        raise HTTPException(status_code=404, detail="Subscription not found")

    if request.plan is not None:
        subscription.plan = request.plan
    if request.status is not None:
        subscription.status = request.status
    if request.trial_start is not None:
        subscription.trial_start = request.trial_start
    if request.trial_end is not None:
        subscription.trial_end = request.trial_end

    subscription.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(subscription)

    return SubscriptionResponse.model_validate(subscription)


# Campaign Endpoints
@app.get("/campaigns")
async def get_campaigns(
    date_preset: str = "last_7d",
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    """Trae campañas REALES de Meta Ads API con métricas."""
    user = await check_subscription(authorization, db)
    config_obj = get_tenant_config(user.id, db)

    if not config_obj.meta_access_token:
        raise HTTPException(status_code=400, detail="Meta API token not configured. Go to Settings to add your Meta credentials.")
    if not config_obj.meta_ad_account_id:
        raise HTTPException(status_code=400, detail="Ad Account ID not configured. Go to Settings to add your Meta Ad Account ID.")

    try:
        from meta_api import get_campaigns as fetch_meta_campaigns
        campaigns = await fetch_meta_campaigns(
            access_token=config_obj.meta_access_token,
            ad_account_id=config_obj.meta_ad_account_id,
            date_preset=date_preset,
        )
        logger.info(f"Fetched {len(campaigns)} campaigns for user {user.id}")
        return campaigns
    except Exception as e:
        logger.error(f"Meta API error for user {user.id}: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=502, detail=f"Error connecting to Meta API: {str(e)}")


@app.post("/campaigns/action")
async def campaign_action(
    request: dict,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    """Pausar o activar una campaña en Meta Ads."""
    user = await check_subscription(authorization, db)
    config_obj = get_tenant_config(user.id, db)

    campaign_id = request.get("campaign_id")
    action = request.get("action", "toggle")

    if not campaign_id:
        raise HTTPException(status_code=400, detail="campaign_id is required")
    if not config_obj.meta_access_token:
        raise HTTPException(status_code=400, detail="Meta API token not configured")

    try:
        from meta_api import pause_campaign, enable_campaign
        if action == "pause":
            success = await pause_campaign(config_obj.meta_access_token, campaign_id)
            new_status = "PAUSED"
        else:
            success = await enable_campaign(config_obj.meta_access_token, campaign_id)
            new_status = "ACTIVE"

        if success:
            return {"campaign_id": campaign_id, "status": new_status, "success": True}
        else:
            raise HTTPException(status_code=502, detail="Meta API returned error")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Campaign action error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=502, detail=f"Error: {str(e)}")


# Agent Endpoints
@app.post("/agent/optimize", response_model=AgentResponse)
async def optimize_campaigns(
    request: AgentRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = await check_subscription(authorization, db)
    config_obj = get_tenant_config(user.id, db)

    if not config_obj.anthropic_api_key:
        raise HTTPException(status_code=400, detail="Anthropic API key not configured")

    try:
        from agents import analyze_campaigns as run_optimizer
        campaigns_data = (request.context or {}).get("campaigns_data", [])

        # Auto-fetch from Meta if no campaigns provided and credentials exist
        if not campaigns_data and config_obj.meta_access_token and config_obj.meta_ad_account_id:
            try:
                from meta_api import get_campaigns as fetch_meta
                campaigns_data = await fetch_meta(
                    access_token=config_obj.meta_access_token,
                    ad_account_id=config_obj.meta_ad_account_id,
                    date_preset="last_7d",
                )
                logger.info(f"Auto-fetched {len(campaigns_data)} campaigns for optimizer agent")
            except Exception as meta_err:
                logger.warning(f"Auto-fetch failed (non-blocking): {meta_err}")

        result = run_optimizer(campaigns_data, config_obj.negocio_info or "", config_obj.anthropic_api_key)

        log_entry = AgentLog(
            user_id=user.id,
            agent_type="optimizer",
            input_summary=request.prompt[:100] if request.prompt else "Campaign analysis",
            output=str(result)[:1000],
        )
        db.add(log_entry)
        db.commit()

        return AgentResponse(result=result, agent="Campaign Optimizer")
    except Exception as e:
        logger.error(f"Optimizer agent error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en agente optimizador: {str(e)}")


@app.post("/agent/finance", response_model=AgentResponse)
async def run_finance_agent(
    request: AgentRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = await check_subscription(authorization, db)
    config_obj = get_tenant_config(user.id, db)

    if not config_obj.anthropic_api_key:
        raise HTTPException(status_code=400, detail="Anthropic API key not configured")

    try:
        from agents import analyze_finances as run_finance
        financial_data = (request.context or {}).get("financial_data", {})
        result = run_finance(financial_data, config_obj.negocio_info or "", config_obj.anthropic_api_key)

        log_entry = AgentLog(
            user_id=user.id, agent_type="finance",
            input_summary=request.prompt[:100] if request.prompt else "Financial analysis",
            output=str(result)[:1000],
        )
        db.add(log_entry)
        db.commit()

        return AgentResponse(result=result, agent="Finance Analyst")
    except Exception as e:
        logger.error(f"Finance agent error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en agente financiero: {str(e)}")


@app.post("/agent/scripts", response_model=AgentResponse)
async def generate_scripts(
    request: AgentRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = await check_subscription(authorization, db)
    config_obj = get_tenant_config(user.id, db)
    
    if not config_obj.anthropic_api_key:
        raise HTTPException(status_code=400, detail="Anthropic API key not configured")
    
    try:
        from agents import generate_scripts as run_scripts
        brief = request.prompt or "Generate ad scripts for current campaigns"
        result = run_scripts(brief, config_obj.negocio_info or "", config_obj.anthropic_api_key)

        log_entry = AgentLog(
            user_id=user.id, agent_type="script_gen",
            input_summary=request.prompt[:100] if request.prompt else "Script generation",
            output=str(result)[:1000],
        )
        db.add(log_entry)
        db.commit()

        return AgentResponse(result=result, agent="Script Generator")
    except Exception as e:
        logger.error(f"Script agent error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en agente de scripts: {str(e)}")


@app.post("/agent/creatives", response_model=AgentResponse)
async def analyze_creatives(
    request: AgentRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = await check_subscription(authorization, db)
    config_obj = get_tenant_config(user.id, db)
    
    if not config_obj.anthropic_api_key:
        raise HTTPException(status_code=400, detail="Anthropic API key not configured")
    
    try:
        from agents import analyze_creatives as run_creatives
        creatives_data = (request.context or {}).get("creatives_data", [])
        result = run_creatives(creatives_data, config_obj.negocio_info or "", config_obj.anthropic_api_key)

        log_entry = AgentLog(
            user_id=user.id, agent_type="creative_director",
            input_summary=request.prompt[:100] if request.prompt else "Creative analysis",
            output=str(result)[:1000],
        )
        db.add(log_entry)
        db.commit()

        return AgentResponse(result=result, agent="Creative Director")
    except Exception as e:
        logger.error(f"Creative agent error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en agente creativo: {str(e)}")


@app.post("/agent/growth", response_model=AgentResponse)
async def get_growth_strategy(
    request: AgentRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = await check_subscription(authorization, db)
    config_obj = get_tenant_config(user.id, db)
    
    if not config_obj.anthropic_api_key:
        raise HTTPException(status_code=400, detail="Anthropic API key not configured")
    
    try:
        from agents import get_growth_strategy as run_growth
        context_data = request.context or {}
        result = run_growth(context_data, config_obj.negocio_info or "", config_obj.anthropic_api_key)

        log_entry = AgentLog(
            user_id=user.id, agent_type="advisor",
            input_summary=request.prompt[:100] if request.prompt else "Growth strategy",
            output=str(result)[:1000],
        )
        db.add(log_entry)
        db.commit()

        return AgentResponse(result=result, agent="Growth Advisor")
    except Exception as e:
        logger.error(f"Growth agent error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en agente de growth: {str(e)}")


@app.post("/agent/cro", response_model=AgentResponse)
async def get_cro_advice(
    request: AgentRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = await check_subscription(authorization, db)
    config_obj = get_tenant_config(user.id, db)
    
    if not config_obj.anthropic_api_key:
        raise HTTPException(status_code=400, detail="Anthropic API key not configured")
    
    try:
        from agents import get_cro_advice as run_cro
        context_data = request.context or {}
        result = run_cro(context_data, config_obj.negocio_info or "", config_obj.anthropic_api_key)

        log_entry = AgentLog(
            user_id=user.id, agent_type="cro",
            input_summary=request.prompt[:100] if request.prompt else "CRO analysis",
            output=str(result)[:1000],
        )
        db.add(log_entry)
        db.commit()

        return AgentResponse(result=result, agent="CRO Advisor")
    except Exception as e:
        logger.error(f"CRO agent error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en agente CRO: {str(e)}")


@app.post("/agent/landing-audit", response_model=AgentResponse)
async def audit_landing_page(
    request: AgentRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = await check_subscription(authorization, db)
    config_obj = get_tenant_config(user.id, db)
    
    if not config_obj.anthropic_api_key:
        raise HTTPException(status_code=400, detail="Anthropic API key not configured")
    
    landing_url = (request.context or {}).get("landing_page_url", "") or config_obj.landing_page_url
    if not landing_url:
        raise HTTPException(status_code=400, detail="Landing page URL not configured")

    try:
        from agents import audit_landing_page as run_landing
        result = run_landing(landing_url, config_obj.negocio_info or "", config_obj.anthropic_api_key)

        log_entry = AgentLog(
            user_id=user.id, agent_type="landing_page_auditor",
            input_summary=landing_url,
            output=str(result)[:1000],
        )
        db.add(log_entry)
        db.commit()

        return AgentResponse(result=result, agent="Landing Page Auditor")
    except Exception as e:
        logger.error(f"Landing audit error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en auditoría de landing: {str(e)}")


@app.post("/agent/full-audit", response_model=AgentResponse)
async def full_audit(
    request: AgentRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = await check_subscription(authorization, db)
    config_obj = get_tenant_config(user.id, db)

    if not config_obj.anthropic_api_key:
        raise HTTPException(status_code=400, detail="Anthropic API key not configured")

    try:
        from agents import run_full_audit as run_orchestrator
        ctx = request.context or {}

        # Gather GA4 summary if configured (optional, non-blocking)
        ga4_summary = ""
        if (hasattr(config_obj, 'ga4_property_id') and hasattr(config_obj, 'ga4_credentials_json') and
            config_obj.ga4_property_id and config_obj.ga4_credentials_json):
            try:
                from agents import get_ga4_summary_for_agents
                ga4_summary = get_ga4_summary_for_agents(
                    config_obj.ga4_property_id,
                    config_obj.ga4_credentials_json,
                )
            except Exception as ga_err:
                logger.warning(f"GA4 data fetch failed (non-blocking): {ga_err}")
                ga4_summary = "[GA4 no disponible]"

        result = run_orchestrator(
            campaigns_data=ctx.get("campaigns_data", []),
            creatives_data=ctx.get("creatives_data", []),
            financial_data=ctx.get("financial_data", {}),
            landing_url=config_obj.landing_page_url or "",
            negocio_info=config_obj.negocio_info or "",
            api_key=config_obj.anthropic_api_key,
        )

        # Append GA4 context to result if available
        if ga4_summary and "[GA4 no disponible]" not in ga4_summary:
            result = result + "\n\n---\n\n" + ga4_summary

        log_entry = AgentLog(
            user_id=user.id, agent_type="orchestrator",
            input_summary="Full business audit (with GA4)" if ga4_summary else "Full business audit",
            output=str(result)[:1000],
        )
        db.add(log_entry)
        db.commit()

        return AgentResponse(result=result, agent="Orchestrator")
    except Exception as e:
        logger.error(f"Orchestrator error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en auditoría completa: {str(e)}")


# ── Google Analytics 4 Endpoints ──

@app.post("/agent/analytics", response_model=AgentResponse)
async def run_analytics_agent(
    request: AgentRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    """Run GA4 analytics analysis agent."""
    user = await check_subscription(authorization, db)
    config_obj = get_tenant_config(user.id, db)

    if not config_obj.anthropic_api_key:
        raise HTTPException(status_code=400, detail="Anthropic API key not configured")
    if not (hasattr(config_obj, 'ga4_property_id') and hasattr(config_obj, 'ga4_credentials_json') and
            config_obj.ga4_property_id and config_obj.ga4_credentials_json):
        raise HTTPException(status_code=400, detail="Google Analytics 4 not configured. Add GA4 Property ID and Service Account credentials in Settings.")

    try:
        from agents import analyze_analytics
        days = (request.context or {}).get("days", 30)
        result = analyze_analytics(
            property_id=config_obj.ga4_property_id,
            ga4_credentials=config_obj.ga4_credentials_json,
            negocio_info=config_obj.negocio_info or "",
            api_key=config_obj.anthropic_api_key,
            days=days,
            landing_url=config_obj.landing_page_url or "",
        )

        log_entry = AgentLog(
            user_id=user.id, agent_type="analytics",
            input_summary=f"GA4 analysis ({days} days)",
            output=str(result)[:1000],
        )
        db.add(log_entry)
        db.commit()

        return AgentResponse(result=result, agent="Analytics Advisor")
    except Exception as e:
        logger.error(f"Analytics agent error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en agente de analytics: {str(e)}")


@app.get("/analytics/data")
async def get_analytics_data(
    days: int = 30,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    """Get raw GA4 data for dashboard display."""
    user = await check_subscription(authorization, db)
    config_obj = get_tenant_config(user.id, db)

    if not (hasattr(config_obj, 'ga4_property_id') and hasattr(config_obj, 'ga4_credentials_json') and
            config_obj.ga4_property_id and config_obj.ga4_credentials_json):
        raise HTTPException(status_code=400, detail="Google Analytics 4 not configured")

    try:
        from agents import fetch_ga4_data
        data = fetch_ga4_data(config_obj.ga4_property_id, config_obj.ga4_credentials_json, days)
        if "error" in data:
            raise HTTPException(status_code=502, detail=f"GA4 API error: {data['error']}")
        return data
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"GA4 data fetch error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching analytics: {str(e)}")


# ── Playbook Nivel Dios Endpoints ──

@app.post("/agent/playbook/copywriter", response_model=AgentResponse)
async def run_playbook_copywriter(
    request: PlaybookCopywriterRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    """Run Playbook Copywriter - Genera copy para landing page, anuncios y email sequence."""
    user = await check_subscription(authorization, db)

    try:
        from agents import run_copywriter_analysis
        result = await run_copywriter_analysis({
            "nicho": request.nicho,
            "audience": request.audience,
            "pain_point": request.pain_point,
            "mechanism_name": request.mechanism_name,
            "mechanism_description": request.mechanism_description,
            "price": request.price,
            "bonos": request.bonos,
            "tone": request.tone,
        })

        log_entry = AgentLog(
            user_id=user.id, agent_type="playbook_copywriter",
            input_summary=f"Copywriter: {request.nicho}",
            output=str(result)[:500],
        )
        db.add(log_entry)
        db.commit()

        return AgentResponse(result=result, agent="Playbook Copywriter")
    except Exception as e:
        logger.error(f"Playbook Copywriter error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en Copywriter: {str(e)}")


@app.post("/agent/playbook/design", response_model=AgentResponse)
async def run_playbook_design(
    request: PlaybookDesignRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    """Run Playbook Design - Genera mockups, auditoría de conversión y estrategia de colores."""
    user = await check_subscription(authorization, db)

    try:
        from agents import run_design_analysis
        result = await run_design_analysis({
            "product_type": request.product_type,
            "product_name": request.product_name,
            "nicho": request.nicho,
            "price": request.price,
            "headline": request.headline,
            "pain_point": request.pain_point,
            "emotion_target": request.emotion_target,
            "bonos": [],
        })

        log_entry = AgentLog(
            user_id=user.id, agent_type="playbook_design",
            input_summary=f"Design: {request.product_name}",
            output=str(result)[:500],
        )
        db.add(log_entry)
        db.commit()

        return AgentResponse(result=result, agent="Playbook Design")
    except Exception as e:
        logger.error(f"Playbook Design error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en Design: {str(e)}")


@app.post("/agent/playbook/social-media", response_model=AgentResponse)
async def run_playbook_social_media(
    request: PlaybookSocialMediaRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    """Run Playbook Social Media - TikTok orgánico + variaciones de creativos para Meta Ads."""
    user = await check_subscription(authorization, db)

    try:
        from agents import run_social_media_analysis
        result = await run_social_media_analysis({
            "nicho": request.nicho,
            "pain_point": request.pain_point,
            "mechanism_name": request.mechanism_name,
            "audience": request.audience,
            "hook": request.hook,
            "angle_type": request.angle_type,
            "format_type": request.format_type,
            "winning_creative_stats": request.winning_creative_stats,
            "current_budget": request.current_budget,
        })

        log_entry = AgentLog(
            user_id=user.id, agent_type="playbook_social_media",
            input_summary=f"Social Media: {request.nicho}",
            output=str(result)[:500],
        )
        db.add(log_entry)
        db.commit()

        return AgentResponse(result=result, agent="Playbook Social Media")
    except Exception as e:
        logger.error(f"Playbook Social Media error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en Social Media: {str(e)}")



# ── Autonomous Actions Endpoints ──

@app.get("/autonomous/actions")
async def get_autonomous_actions(
    limit: int = 50,
    status: Optional[str] = None,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    """Get history of autonomous actions (last 50 by default)."""
    user = await check_subscription(authorization, db)

    query = db.query(AutonomousActionLog).filter(AutonomousActionLog.user_id == user.id)
    if status:
        query = query.filter(AutonomousActionLog.status == status)

    actions = query.order_by(AutonomousActionLog.created_at.desc()).limit(limit).all()

    return [
        {
            "id": a.id,
            "action_type": a.action_type,
            "status": a.status,
            "target": a.target,
            "description": a.description,
            "triggered_by": a.triggered_by,
            "requires_approval": a.requires_approval,
            "approved_by": a.approved_by,
            "created_at": a.created_at.isoformat(),
            "executed_at": a.executed_at.isoformat() if a.executed_at else None,
        }
        for a in actions
    ]


@app.post("/autonomous/actions/{action_id}/approve")
async def approve_action(
    action_id: int,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    """Admin approves an autonomous action (e.g., scale budget)."""
    user = await check_subscription(authorization, db)

    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can approve actions")

    action = db.query(AutonomousActionLog).filter(AutonomousActionLog.id == action_id).first()
    if not action:
        raise HTTPException(status_code=404, detail="Action not found")

    if action.status != "pending":
        raise HTTPException(status_code=400, detail=f"Action is already {action.status}")

    # Mark as approved
    action.status = "approved"
    action.approved_by = str(user.id)
    action.approved_at = datetime.utcnow()

    # TODO: Here you would actually execute the action in Meta Ads API
    # For now, just mark as approved
    # action.status = "executed"
    # action.executed_at = datetime.utcnow()

    db.commit()

    return {"status": "approved", "action_id": action_id, "approved_at": action.approved_at}


@app.post("/autonomous/actions/{action_id}/reject")
async def reject_action(
    action_id: int,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    """Admin rejects an autonomous action."""
    user = await check_subscription(authorization, db)

    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can reject actions")

    action = db.query(AutonomousActionLog).filter(AutonomousActionLog.id == action_id).first()
    if not action:
        raise HTTPException(status_code=404, detail="Action not found")

    if action.status != "pending":
        raise HTTPException(status_code=400, detail=f"Action is already {action.status}")

    action.status = "cancelled"
    action.approved_by = str(user.id)
    action.approved_at = datetime.utcnow()
    db.commit()

    return {"status": "rejected", "action_id": action_id}


@app.post("/agent/run", response_model=AgentResponse)
async def run_agent(
    request: AgentRunRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = await check_subscription(authorization, db)
    config_obj = get_tenant_config(user.id, db)

    if not config_obj.anthropic_api_key:
        raise HTTPException(status_code=400, detail="Anthropic API key not configured")

    agent_type = request.agent_type.lower()

    # Map agent types to their logic
    agent_handlers = {
        "optimizer": "Campaign Optimizer",
        "finance": "Finance Analyst",
        "script_gen": "Script Generator",
        "creative_director": "Creative Director",
        "advisor": "Growth Advisor",
        "cro": "CRO Advisor",
        "landing_page_auditor": "Landing Page Auditor",
        "orchestrator": "Orchestrator",
    }

    if agent_type not in agent_handlers:
        raise HTTPException(status_code=400, detail=f"Unknown agent type: {agent_type}")

    try:
        # Real agent dispatch — no placeholders
        from agents import (
            analyze_campaigns, analyze_finances, generate_scripts,
            analyze_creatives, get_growth_strategy, get_cro_advice,
            audit_landing_page, run_full_audit,
        )

        api_key = config_obj.anthropic_api_key
        negocio = config_obj.negocio_info or ""
        ctx = request.context or {}

        if agent_type == "optimizer":
            campaigns_data = ctx.get("campaigns_data", [])
            result = analyze_campaigns(campaigns_data, negocio, api_key)
        elif agent_type == "finance":
            financial_data = ctx.get("financial_data", {})
            result = analyze_finances(financial_data, negocio, api_key)
        elif agent_type == "script_gen":
            result = generate_scripts(request.input or "Generate ad scripts", negocio, api_key)
        elif agent_type == "creative_director":
            creatives_data = ctx.get("creatives_data", [])
            result = analyze_creatives(creatives_data, negocio, api_key)
        elif agent_type == "advisor":
            result = get_growth_strategy(ctx, negocio, api_key)
        elif agent_type == "cro":
            result = get_cro_advice(ctx, negocio, api_key)
        elif agent_type == "landing_page_auditor":
            landing_url = ctx.get("landing_page_url", "") or config_obj.landing_page_url or ""
            result = audit_landing_page(landing_url, negocio, api_key)
        elif agent_type == "orchestrator":
            result = run_full_audit(
                campaigns_data=ctx.get("campaigns_data", []),
                creatives_data=ctx.get("creatives_data", []),
                financial_data=ctx.get("financial_data", {}),
                landing_url=config_obj.landing_page_url or "",
                negocio_info=negocio,
                api_key=api_key,
            )
        else:
            result = f"Agent '{agent_type}' not implemented"

        # Log the result
        log_entry = AgentLog(
            user_id=user.id,
            agent_type=agent_type,
            input_summary=request.input[:100] if request.input else "Agent execution",
            output=str(result)[:1000],
        )
        db.add(log_entry)
        db.commit()

        return AgentResponse(
            result=result,
            agent=agent_handlers[agent_type],
        )
    except Exception as e:
        logger.error(f"Agent run error ({agent_type}): {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error in {agent_type} agent: {str(e)}")


# Finance Endpoints
@app.post("/finance/upload")
async def upload_financial_data(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = await check_subscription(authorization, db)
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        records_created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            record = FinancialRecord(
                user_id=user.id,
                periodo=str(row[0]) if row[0] else "",
                ingresos=int(row[1]) if row[1] else None,
                costos=int(row[2]) if row[2] else None,
                ad_spend=int(row[3]) if row[3] else None,
                devoluciones=int(row[4]) if row[4] else None,
                ordenes=int(row[5]) if row[5] else None,
            )
            db.add(record)
            records_created += 1
        
        db.commit()
        return {
            "records_created": records_created,
            "message": "Financial data uploaded successfully",
        }
    except Exception as e:
        logger.error(f"Error processing financial data: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=400, detail="Invalid request data")


@app.get("/finance/records", response_model=List[FinancialRecordResponse])
async def get_financial_records(
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = await check_subscription(authorization, db)

    records = db.query(FinancialRecord).filter(
        FinancialRecord.user_id == user.id,
    ).order_by(FinancialRecord.created_at.desc()).all()

    return [FinancialRecordResponse.model_validate(record) for record in records]


@app.get("/financials", response_model=List[FinancialRecordResponse])
async def get_financials(
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = await check_subscription(authorization, db)

    records = db.query(FinancialRecord).filter(
        FinancialRecord.user_id == user.id,
    ).order_by(FinancialRecord.created_at.desc()).all()

    return [FinancialRecordResponse.model_validate(record) for record in records]


@app.post("/financials", response_model=FinancialRecordResponse)
async def create_financial_record(
    request: FinancialRecordCreate,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = await check_subscription(authorization, db)

    record = FinancialRecord(
        user_id=user.id,
        periodo=request.periodo,
        ingresos=request.ingresos,
        costos=request.costos,
        ad_spend=request.ad_spend,
        devoluciones=request.devoluciones,
        ordenes=request.ordenes,
    )
    db.add(record)
    db.commit()
    db.refresh(record)

    return FinancialRecordResponse.model_validate(record)


@app.post("/financials/upload")
async def upload_financial_records(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = await check_subscription(authorization, db)

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        records_created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            record = FinancialRecord(
                user_id=user.id,
                periodo=str(row[0]) if row[0] else "",
                ingresos=int(row[1]) if row[1] else None,
                costos=int(row[2]) if row[2] else None,
                ad_spend=int(row[3]) if row[3] else None,
                devoluciones=int(row[4]) if row[4] else None,
                ordenes=int(row[5]) if row[5] else None,
            )
            db.add(record)
            records_created += 1

        db.commit()
        return {
            "records_created": records_created,
            "message": "Financial data uploaded successfully",
        }
    except Exception as e:
        logger.error(f"Error processing financial data: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=400, detail="Invalid request data")


# Shopify Endpoints
@app.post("/orders/webhook")
async def shopify_webhook(
    request: dict,
    db: Session = Depends(get_db),
):
    # Webhook from Shopify - need to identify user by store URL
    # For now, return success
    return {"status": "received"}


@app.get("/orders", response_model=List[ShopifyOrderResponse])
async def get_orders(
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = await check_subscription(authorization, db)
    
    orders = db.query(ShopifyOrder).filter(
        ShopifyOrder.user_id == user.id,
    ).order_by(ShopifyOrder.created_at.desc()).all()
    
    return [ShopifyOrderResponse.model_validate(order) for order in orders]


# ── Shared Memory + Multi-Agent Endpoints ──

class MemoryWriteRequest(BaseModel):
    value: Any
    memory_type: str = "product"


class InfoproductoRunRequest(BaseModel):
    step_id: str
    state: Dict[str, Any]


class TikTokAdsRunRequest(BaseModel):
    mode: str  # "strategy" | "optimize"
    payload: Dict[str, Any] = {}


class TikTokVideoRequest(BaseModel):
    context: Dict[str, Any] = {}


@app.get("/memory/{key}")
async def read_memory(
    key: str,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = get_current_user_from_header(authorization, db)
    from agents.shared_memory import SharedMemoryStore
    memory = SharedMemoryStore(db, user.id)
    value = memory.read(key)
    return {"key": key, "value": value}


@app.post("/memory/{key}")
async def write_memory(
    key: str,
    request: MemoryWriteRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = get_current_user_from_header(authorization, db)
    from agents.shared_memory import SharedMemoryStore
    memory = SharedMemoryStore(db, user.id)
    memory.write(key, request.value, agent="user", memory_type=request.memory_type)
    return {"key": key, "ok": True}


@app.post("/agents/infoproducto/run")
async def run_infoproducto_endpoint(
    request: InfoproductoRunRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = get_current_user_from_header(authorization, db)
    config_obj = get_tenant_config(user.id, db)
    if not config_obj.anthropic_api_key:
        raise HTTPException(status_code=400, detail="Anthropic API key not configured")
    try:
        from agents.infoproducto_agent import run_infoproducto_step
        output = run_infoproducto_step(
            db=db,
            user_id=user.id,
            step_id=request.step_id,
            state=request.state,
            api_key=config_obj.anthropic_api_key,
        )
        return {"step_id": request.step_id, "output": output}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Infoproducto agent error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en agente infoproducto: {str(e)}")


@app.post("/agents/tiktok-ads/run")
async def run_tiktok_ads_endpoint(
    request: TikTokAdsRunRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = get_current_user_from_header(authorization, db)
    config_obj = get_tenant_config(user.id, db)
    if not config_obj.anthropic_api_key:
        raise HTTPException(status_code=400, detail="Anthropic API key not configured")
    try:
        from agents.tiktok_ads_agent import run_tiktok_ads
        output = run_tiktok_ads(
            db=db,
            user_id=user.id,
            mode=request.mode,
            payload=request.payload,
            api_key=config_obj.anthropic_api_key,
        )
        return {"mode": request.mode, "output": output}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"TikTok Ads agent error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en agente TikTok Ads: {str(e)}")


@app.post("/agents/tiktok/video")
async def generate_tiktok_video_endpoint(
    request: TikTokVideoRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = get_current_user_from_header(authorization, db)
    config_obj = get_tenant_config(user.id, db)
    if not config_obj.anthropic_api_key:
        raise HTTPException(status_code=400, detail="Anthropic API key not configured")
    try:
        from agents.tiktok_creator import TikTokCreatorAgent
        from agents.shared_memory import SharedMemoryStore
        memory = SharedMemoryStore(db, user.id)
        ctx = {**memory.get_full_context(), **request.context}
        agent = TikTokCreatorAgent(api_key=config_obj.anthropic_api_key)
        result = agent.generate_daily_video(ctx)
        return result
    except Exception as e:
        logger.error(f"TikTok video agent error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en agente TikTok orgánico: {str(e)}")


@app.post("/agents/tiktok/calendar")
async def generate_tiktok_calendar_endpoint(
    request: TikTokVideoRequest,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    user = get_current_user_from_header(authorization, db)
    config_obj = get_tenant_config(user.id, db)
    if not config_obj.anthropic_api_key:
        raise HTTPException(status_code=400, detail="Anthropic API key not configured")
    try:
        from agents.tiktok_creator import TikTokCreatorAgent
        from agents.shared_memory import SharedMemoryStore
        memory = SharedMemoryStore(db, user.id)
        ctx = {**memory.get_full_context(), **request.context}
        agent = TikTokCreatorAgent(api_key=config_obj.anthropic_api_key)
        result = agent.generate_weekly_calendar(ctx)
        return result
    except Exception as e:
        logger.error(f"TikTok calendar agent error: {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error en agente calendario TikTok: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
